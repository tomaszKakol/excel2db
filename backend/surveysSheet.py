import openpyxl
import xlsxwriter
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from backend import baseDb as base, labels
from frontend import charts as data
import statistics

# kolumny podstawowe
# cols_base = ["Prowadzący", "Nazwa przedmiotu", "Rok ankiety", "Semestr"]

# instrukcje zawarte w tym skrypcie służą do utworzenia raportu, tj. utworzenia akrusza kalkulacyjnego,
# w którym zostaną zawarte najważniejsze informacje z przeprowadzonych ankiet
def create_menu(db, cols_base):
    workbook = xlsxwriter.Workbook(labels.SHEET_NAME)
    # każdy wydział będzie zawarty na nowej zakładce w pliku .xlsx
    columns = '"%s", "%s"' % (labels.ID, labels.DEPARTMENT_NAME,)
    departmentsData = base.select_table_data_1(db, labels.TABLE_DEPARTMENTS, columns)

    # departments = []
    # departmentsNo = None
    if departmentsData:
        # print(departmentsData)
        departmentsNo = len(departmentsData)
        for departmentNo in range(departmentsNo):
            worksheet = workbook.add_worksheet((departmentsData[departmentNo])[1])
            # kolumny podstawowe
            # cols_base = ["Prowadzący", "Nazwa przedmiotu", "Rok ankiety", "Semestr"]

            # kolumny zmienne, dotyczące pytań z ankiet, zawartych w tabeli 'wyniki' w naszej bazie danych
            results_cols = base.select_table_columns(db, labels.TABLE_RESULTS)
            # print(results_cols, len(results_cols))

            cols_answers = []
            for i in range(len(results_cols)):
                if "A_Pytanie" in results_cols[i]:
                    # cols_base.append(results_cols[i])
                    cols_answers.append(results_cols[i])

            cell_format = workbook.add_format({'align': 'center',
                                               'valign': 'vcenter',
                                               'border': 1})
            cell_format.set_pattern(1)  # This is optional when using a solid fill.
            cell_format.set_bg_color('green')
            blue = workbook.add_format({'color': 'blue'})

            for x in range(len(cols_base)):
                column_letter_start = get_column_letter(x+1)
                worksheet.merge_range('%s1:%s2' % (column_letter_start, column_letter_start), cols_base[x],
                                      cell_format)

            # dla kolumn typu 'Pytanie_1' będą 4 podkolumny dla różnych statystyk (mediana, średnia ...)
            # taki efekt uzyskamy przez scalenie komórek w arkuszu
            for colsCounter in range(len(cols_answers)):
                column_letter_start = get_column_letter(len(cols_base) + 1 + colsCounter * 4)
                column_letter_end = get_column_letter(len(cols_base) + 4 + colsCounter * 4)
                # print('%s1:%s1' % (column_letter_start, column_letter_end))
                worksheet.merge_range('%s1:%s1' % (column_letter_start, column_letter_end), cols_answers[colsCounter], cell_format)
                worksheet.write_rich_string('%s1' % (column_letter_start, ),
                                            ' ',
                                            blue, cols_answers[colsCounter],
                                            cell_format)

                worksheet.write(1, len(cols_base) + 0 + colsCounter * 4, 'Liczba odpowiedzi', cell_format)
                worksheet.write(1, len(cols_base) + 1 + colsCounter * 4, 'Mediana', cell_format)
                worksheet.write(1, len(cols_base) + 2 + colsCounter * 4, 'Średnia', cell_format)
                worksheet.write(1, len(cols_base) + 3 + colsCounter * 4, 'Odchylenie standardowe', cell_format)

    workbook.close()


def fill_sheets(db, cols_base):
    wb = openpyxl.load_workbook(labels.SHEET_NAME)
    sheet = wb.active
    yellowFill = PatternFill(start_color='FFFF33',
                          end_color='FFFF33',
                          fill_type='solid')

    # create sheets as departments names
    columns = '"%s", "%s"' % (labels.ID, labels.DEPARTMENT_NAME,)
    departmentsData = base.select_table_data_1(db, labels.TABLE_DEPARTMENTS, columns)

    #po otwarciu arkusza przechodzimy do pierwszej zakładki wydziału
    departments = []
    departmentsNo = None
    if departmentsData:
        # print(departmentsData)
        departmentsNo = len(departmentsData)
        if departmentsNo > 1:
            # wb_sheet = wb['Sheet1']
            # wb_sheet.title = (departmentsData[0])[1]
            departments.append((departmentsData[0])[1])
        for i in range(departmentsNo - 1):
            # wb.create_sheet((departmentsData[i+1])[1])
            departments.append((departmentsData[i+1])[1])

    # --- USERS ---
    if departments:
        columns2 = '%s, "%s"' % (labels.ID, labels.USER_NAME,)
        for departmentNo in range(departmentsNo):
            # zaczynamy wypełniać tabelę od 3 wiersza, bo 2 pierwsze zajmują nagłówki kolumn
            courseRowIdx = 3
            sheet = wb.get_sheet_by_name(departments[departmentNo])

            # cols_base = ["Prowadzący", "Nazwa przedmiotu", "Rok ankiety", "Semestr"]
            results_cols = base.select_table_columns(db, labels.TABLE_RESULTS)

            cols_answers = []
            for i in range(len(results_cols)):
                if "A_Pytanie" in results_cols[i]:
                    cols_base.append(results_cols[i])
                    cols_answers.append(results_cols[i])
            # sheet.append(columnsNames)

            departmentId = (departmentsData[departmentNo])[0]
            condition2 = '"%s" = %s' % (labels.DEPARTMENT_ID, departmentId,)
            usersData = base.select_table_data_2(db, labels.TABLE_USERS, columns2, condition2)
            # print(usersData)

            usersCellsIdx = []
            if usersData:
                for i, user in enumerate(usersData):
                    user = list(user)
                    userId = user[0]

                    # sheet.append([user[1]])
                    # gdybyśmy chcieli scalić komórki, to musimy zachować ostatni wiersz danego prowadzącego
                    # print('komórka na wpisanie nazwy uzytkownika:', 'A%d' % (sheet.max_row,))
                    usersCellsIdx.append(sheet.max_row)


                    columns3 = '"%s", "%s", "%s", "%s"' % (labels.ID, labels.COURSE_NAME, labels.YEAR, labels.TERM)
                    condition3 = '"%s" = %s and "%s" = %s' % (
                        labels.USER_ID, user[0], labels.DEPARTMENT_ID, (departmentsData[departmentNo])[0],)
                    coursesData = base.select_table_data_2(db, labels.TABLE_COURSES, columns3, condition3)

                    for courseCounter, course in enumerate(coursesData):
                        # print(course)

                        col = "A" + str(courseRowIdx)
                        sheet[col] = user[1]

                        courseId = (coursesData[courseCounter])[0]
                        col = "B" + str(courseRowIdx)
                        sheet[col] = (coursesData[courseCounter])[1]

                        col = "C" + str(courseRowIdx)
                        sheet[col] = (coursesData[courseCounter])[2]

                        col = "D" + str(courseRowIdx)
                        sheet[col] = (coursesData[courseCounter])[3]

                        # wyciągamy nazwy (litery) istniejących kolumn
                        dims = {}
                        columnsLetters = []
                        for row in sheet.rows:
                            for cell in row:
                                # if cell.value:
                                #     dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
                                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
                        for col, value in dims.items():
                            column_letter = get_column_letter(col)
                            columnsLetters.append(column_letter)

                        # print(userId, departmentId, courseId, cols_answers)
                        values, stat_cols, filtered, answersNo = data.retrive_data2(
                            db, labels.TABLE_RESULTS, userId, departmentId, courseId, cols_answers)

                        # wyznaczamy wartości na podstawie danych zwróconych z bazy danych
                        median = []
                        average = []
                        stdev = []
                        # print(values, stat_cols, filtered, answersNo)
                        if values and stat_cols and filtered and answersNo:
                            for filteredCounter in range(len(filtered)):
                                if filtered[filteredCounter]:
                                    median.append(statistics.median(filtered[filteredCounter]))
                                    average.append(statistics.mean(filtered[filteredCounter]))
                                    stdev.append(statistics.stdev(filtered[filteredCounter]))
                                else:
                                    median.append(None)
                                    average.append(None)

                            #wyznaczamy odpowiednie komórki w arkuszu, zgodnie z menu i wpisujemy tam obliczone wartości
                            y = 1
                            for x in range(len(cols_answers)):
                                y = y + 3
                                if x < len(answersNo):
                                    ansNo = answersNo[x]
                                    m = median[x]
                                    av = average[x]
                                    sd = stdev[x]

                                    col = str(columnsLetters[x + y]) + str(courseRowIdx)
                                    sheet[col].alignment = Alignment(horizontal="right")
                                    if isinstance(ansNo, int):
                                        sheet[col] = '%d' % (ansNo,)
                                    elif isinstance(ansNo, float):
                                        sheet[col] = '%.1f' % (ansNo,)
                                    else:
                                        sheet[col] = ansNo
                                    col = str(columnsLetters[x+y+1]) + str(courseRowIdx)
                                    sheet[col].alignment = Alignment(horizontal="right")
                                    if isinstance(m, float):
                                        sheet[col] = '%.1f  ' % (m,)
                                    else:
                                        sheet[col] = m
                                    col = str(columnsLetters[x+y+2]) + str(courseRowIdx)
                                    sheet[col].alignment = Alignment(horizontal="right")
                                    if isinstance(av, float):
                                        sheet[col] = '%.2f  ' % (av,)
                                    else:
                                        sheet[col] = av
                                    col = str(columnsLetters[x+y+3]) + str(courseRowIdx)
                                    sheet[col].alignment = Alignment(horizontal="right")
                                    if isinstance(sd, float):
                                        sheet[col] = '%.4f  ' % (sd,)
                                    else:
                                        sheet[col] = sd

                        courseRowIdx = courseRowIdx + 1

                # rozciągamy szerokość kolumn zgodnie z zawartością komórek w tych kolumnach,
                # tak aby wszystko było widać w arkuszu
                ws = sheet
                dims = {}
                for row in ws.rows:
                    for cell in row:
                        if cell.value:
                            dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
                for col, value in dims.items():
                    column_letter = get_column_letter(col)
                    # print(column_letter)
                    ws.column_dimensions[column_letter].width = value * 1.2

                row_count = sheet.max_row
                # column_count = sheet.max_column

                thin_border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))
                for row in range(row_count + 1):
                    if row > 2:
                        for idx in range(len(columnsLetters)):
                            if idx < 4:
                                # print('%s%d' % (columnsLetters[idx], row))
                                sheet['%s%d' % (columnsLetters[idx], row)].fill = yellowFill
                                sheet['%s%d' % (columnsLetters[idx], row)].border = thin_border
                                # sheet.merge_cells(start_row=3, start_column=1, end_row=8, end_column=1)

                #  scalenie komórek w pierwszej kolumnie (prowadzący)
                # # print(len(usersCellsIdx))
                # for idx in range(len(usersCellsIdx)):
                #     if idx + 1 == len(usersCellsIdx):
                #         sheet.merge_cells(start_row=usersCellsIdx[idx], start_column=1,
                #                           end_row=courseRowIdx - 1, end_column=1)
                #     else:
                #         sheet.merge_cells(start_row=usersCellsIdx[idx], start_column=1,
                #                               end_row=usersCellsIdx[idx + 1] - 1, end_column=1)


    wb.save(labels.SHEET_NAME)


